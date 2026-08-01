import datetime as dt
from io import BytesIO

from openpyxl import Workbook, load_workbook

SHEET_COLUMNS = {
    "historial_fondo": ["fecha", "valor_total", "precio_cuota", "cuotas_circ", "trm"],
    "movimientos": ["fecha", "persona", "tipo", "monto", "precio_cuota_dia", "cuotas", "monto_cop", "trm_dia"],
    "participantes_config": ["fecha", "nombre", "accion"],
}

TEXT_COLUMNS = {"fecha", "persona", "tipo", "nombre", "accion"}

# Columnas sin las cuales la fila no significa nada: si falta alguna, todas las filas de esa
# hoja entrarían con "" o 0.0 y el import es destructivo — se aborta en vez de reemplazar la
# DB por filas vacías. Las que no están acá (trm, monto_cop, precio_cuota_dia) sí pueden
# faltar: se derivan o solo afectan la vista en COP.
REQUIRED_COLUMNS = {
    "historial_fondo": ["fecha", "valor_total", "precio_cuota", "cuotas_circ"],
    "movimientos": ["fecha", "persona", "tipo", "monto", "cuotas"],
    "participantes_config": ["fecha", "nombre", "accion"],
}

ENUM_COLUMNS = {
    "tipo": {"aporte", "retiro"},
    "accion": {"agregar", "quitar"},
}

# Acepta tanto nuestro propio formato de export (columnas = nombres de la tabla) como los
# nombres de columna del Google Sheet original que reemplaza esta DB (p.ej. "valor_total_usd",
# "valor_pesos") — así el mismo importador sirve tanto para restaurar un backup como para
# migrar el histórico real del Sheet.
COLUMN_ALIASES = {
    "historial_fondo": {
        "fecha": ["fecha"],
        "valor_total": ["valor_total", "valor_total_usd"],
        "precio_cuota": ["precio_cuota", "precio_cuota_usd"],
        "cuotas_circ": ["cuotas_circ", "cuotas_en_circulacion"],
        "trm": ["trm"],
    },
    "movimientos": {
        "fecha": ["fecha"],
        "persona": ["persona"],
        "tipo": ["tipo"],
        "monto": ["monto", "monto_usd"],
        "precio_cuota_dia": ["precio_cuota_dia"],
        "cuotas": ["cuotas"],
        "monto_cop": ["monto_cop", "valor_pesos"],
        "trm_dia": ["trm_dia", "trm"],
    },
    "participantes_config": {
        "fecha": ["fecha"],
        "nombre": ["nombre"],
        "accion": ["accion"],
    },
}


def build_workbook(historial, movimientos, participantes_config):
    rows_by_sheet = {
        "historial_fondo": historial,
        "movimientos": movimientos,
        "participantes_config": participantes_config,
    }
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, cols in SHEET_COLUMNS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(cols)
        for row in rows_by_sheet[sheet_name]:
            ws.append([row[c] for c in cols])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _to_fecha_str(value):
    if isinstance(value, dt.datetime):
        if (value.hour, value.minute, value.second) == (0, 0, 0):
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%dT%H:%M")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


class InvalidWorkbook(Exception):
    pass


# El import reemplaza toda la DB: una celda con "aporte " o "Aporte" se normaliza, pero
# cualquier otra cosa se rechaza en vez de entrar y romper la matemática de cuotas en silencio.
def _to_enum(value, col, sheet_name, fila):
    valor = str(value).strip().lower() if value is not None else ""
    if valor not in ENUM_COLUMNS[col]:
        opciones = " / ".join(sorted(ENUM_COLUMNS[col]))
        raise InvalidWorkbook(
            f"{sheet_name} fila {fila}: '{col}' es '{value}' y solo acepta {opciones}"
        )
    return valor


def _to_float(value, col, sheet_name, fila):
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        raise InvalidWorkbook(f"{sheet_name} fila {fila}: '{col}' no es un número ('{value}')") from None


def parse_workbook(content: bytes):
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise InvalidWorkbook("Archivo xlsx inválido") from exc

    parsed = {}
    for sheet_name, cols in SHEET_COLUMNS.items():
        parsed[sheet_name] = []
        if sheet_name not in wb.sheetnames:
            continue
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
        header_set = set(header)
        aliases = COLUMN_ALIASES[sheet_name]
        # Para cada columna canónica, qué nombre de columna real trae este archivo (si trae alguno).
        source = {col: next((a for a in alts if a in header_set), None) for col, alts in aliases.items()}

        faltan = [c for c in REQUIRED_COLUMNS[sheet_name] if source[c] is None]
        if faltan:
            raise InvalidWorkbook(
                f"{sheet_name}: faltan columnas requeridas ({', '.join(faltan)})"
            )

        for n, raw in enumerate(rows[1:], start=2):
            if raw is None or all(v is None for v in raw):
                continue
            record = dict(zip(header, raw))
            if record.get(source["fecha"]) is None:
                continue
            item = {}
            for col in cols:
                value = record.get(source[col]) if source[col] else None
                if col == "fecha":
                    item[col] = _to_fecha_str(value)
                elif col in ENUM_COLUMNS:
                    item[col] = _to_enum(value, col, sheet_name, n)
                elif col in TEXT_COLUMNS:
                    item[col] = str(value).strip() if value is not None else ""
                else:
                    item[col] = _to_float(value, col, sheet_name, n)
            parsed[sheet_name].append(item)
    return parsed
