import importlib
import io
import sqlite3

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("DB_PATH", str(tmp_path / "test.db"))
    monkeypatch.setenv("ADMIN_PASSWORD", "s3cret")

    from app import db as db_module
    from app import main as main_module

    importlib.reload(db_module)
    importlib.reload(main_module)

    with TestClient(main_module.app) as c:
        yield c


def test_health(client):
    r = client.get("/api/health")
    assert r.status_code == 200
    assert r.json() == {"ok": True}


def test_admin_password_defaults_to_admin_when_unset(tmp_path, monkeypatch):
    monkeypatch.setenv("DB_PATH", str(tmp_path / "test.db"))
    monkeypatch.delenv("ADMIN_PASSWORD", raising=False)

    from app import db as db_module
    from app import main as main_module

    importlib.reload(db_module)
    importlib.reload(main_module)

    with TestClient(main_module.app) as c:
        r = c.post("/api/auth/verify", headers={"X-Admin-Key": "admin"})
        assert r.status_code == 200

        r = c.post("/api/auth/verify", headers={"X-Admin-Key": "not-admin"})
        assert r.status_code == 401


def test_get_all_empty(client):
    r = client.get("/api/all")
    assert r.status_code == 200
    assert r.json() == {"historial": [], "movimientos": [], "participantes_config": []}


def test_post_movimiento_requires_auth(client):
    payload = {
        "fecha": "2026-01-10T00:00", "persona": "Patico", "tipo": "aporte",
        "monto_usd": 100, "precio_cuota_dia": 1.0, "cuotas": 100, "monto_cop": 330000, "trm_dia": 3300,
    }
    r = client.post("/api/movimiento", json=payload)
    assert r.status_code == 401

    r = client.post("/api/movimiento", json=payload, headers={"X-Admin-Key": "wrong"})
    assert r.status_code == 401


def test_post_movimiento_then_get_all(client):
    payload = {
        "fecha": "2026-01-10T00:00", "persona": "Patico", "tipo": "aporte",
        "monto_usd": 100, "precio_cuota_dia": 1.0, "cuotas": 100, "monto_cop": 330000, "trm_dia": 3300,
    }
    r = client.post("/api/movimiento", json=payload, headers={"X-Admin-Key": "s3cret"})
    assert r.status_code == 201

    r = client.get("/api/all")
    assert len(r.json()["movimientos"]) == 1
    assert r.json()["movimientos"][0]["persona"] == "Patico"


def test_post_fondo(client):
    payload = {
        "fecha": "2026-01-10T00:00", "valor_total_usd": 400, "precio_cuota_usd": 1.0,
        "cuotas_en_circulacion": 400, "trm": 3300,
    }
    r = client.post("/api/fondo", json=payload, headers={"X-Admin-Key": "s3cret"})
    assert r.status_code == 201

    r = client.get("/api/all")
    assert len(r.json()["historial"]) == 1


def test_auth_verify(client):
    r = client.post("/api/auth/verify")
    assert r.status_code == 401

    r = client.post("/api/auth/verify", headers={"X-Admin-Key": "s3cret"})
    assert r.status_code == 200


def test_post_participante(client):
    payload = {"fecha": "2026-01-10T00:00", "nombre": "Patico", "accion": "agregar"}
    r = client.post("/api/participante", json=payload, headers={"X-Admin-Key": "s3cret"})
    assert r.status_code == 201

    r = client.get("/api/all")
    assert r.json()["participantes_config"] == [payload]


def test_export_xlsx(client):
    client.post("/api/movimiento", headers={"X-Admin-Key": "s3cret"}, json={
        "fecha": "2026-01-10T00:00", "persona": "Patico", "tipo": "aporte",
        "monto_usd": 100, "precio_cuota_dia": 1.0, "cuotas": 100, "monto_cop": 330000, "trm_dia": 3300,
    })

    r = client.get("/api/export")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["historial_fondo", "movimientos", "participantes_config"]
    rows = list(wb["movimientos"].iter_rows(values_only=True))
    assert rows[0] == ("fecha", "persona", "tipo", "monto", "precio_cuota_dia", "cuotas", "monto_cop", "trm_dia")
    assert rows[1][1] == "Patico"


def _xlsx_bytes(sheets):
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def test_import_requires_auth(client):
    content = _xlsx_bytes({"movimientos": [["fecha", "persona", "tipo", "monto", "precio_cuota_dia", "cuotas", "monto_cop", "trm_dia"]]})
    r = client.post("/api/import", files={"file": ("data.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 401


def test_import_replaces_existing_data(client):
    client.post("/api/participante", headers={"X-Admin-Key": "s3cret"},
                json={"fecha": "2026-01-01T00:00", "nombre": "Viejo", "accion": "agregar"})

    content = _xlsx_bytes({
        "historial_fondo": [
            ["fecha", "valor_total", "precio_cuota", "cuotas_circ", "trm"],
            ["2026-01-10", 400, 1.0, 400, 3300],
        ],
        "movimientos": [
            ["fecha", "persona", "tipo", "monto", "precio_cuota_dia", "cuotas", "monto_cop", "trm_dia"],
            ["2026-01-10T00:00", "Patico", "aporte", 100, 1.0, 100, 330000, 3300],
        ],
        "participantes_config": [
            ["fecha", "nombre", "accion"],
            ["2026-01-10T00:00", "Patico", "agregar"],
        ],
    })

    r = client.post("/api/import", headers={"X-Admin-Key": "s3cret"},
                     files={"file": ("data.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200
    assert r.json()["counts"] == {"historial_fondo": 1, "movimientos": 1, "participantes_config": 1}

    data = client.get("/api/all").json()
    assert len(data["historial"]) == 1
    assert len(data["movimientos"]) == 1
    assert data["participantes_config"] == [{"fecha": "2026-01-10T00:00", "nombre": "Patico", "accion": "agregar"}]


def test_import_invalid_file(client):
    r = client.post("/api/import", headers={"X-Admin-Key": "s3cret"},
                     files={"file": ("data.xlsx", b"not an xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 400


def test_auth_con_clave_no_ascii_da_401_no_500(client):
    # El browser manda los valores de header en latin-1; Starlette los decodifica a str y
    # secrets.compare_digest sobre str no-ASCII tiraba TypeError → 500 en vez de 401.
    r = client.post("/api/auth/verify", headers={"X-Admin-Key": "clavé".encode("latin-1")})
    assert r.status_code == 401


def test_movimiento_con_fondo_es_atomico(client):
    payload = {
        "fecha": "2026-01-10T00:00", "persona": "Patico", "tipo": "aporte",
        "monto_usd": 100, "precio_cuota_dia": 1.0, "cuotas": 100, "monto_cop": 330000, "trm_dia": 3300,
        "fondo": {
            "fecha": "2026-01-10T00:00", "valor_total_usd": 100, "precio_cuota_usd": 1.0,
            "cuotas_en_circulacion": 100, "trm": 3300,
        },
    }
    r = client.post("/api/movimiento", json=payload, headers={"X-Admin-Key": "s3cret"})
    assert r.status_code == 201

    data = client.get("/api/all").json()
    assert len(data["movimientos"]) == 1
    assert len(data["historial"]) == 1


def test_movimiento_con_fondo_invalido_no_guarda_nada(client):
    payload = {
        "fecha": "2026-01-10T00:00", "persona": "Patico", "tipo": "aporte",
        "monto_usd": 100, "precio_cuota_dia": 1.0, "cuotas": 100, "monto_cop": 330000, "trm_dia": 3300,
        "fondo": {
            "fecha": "no-es-fecha", "valor_total_usd": 100, "precio_cuota_usd": 1.0,
            "cuotas_en_circulacion": 100, "trm": 3300,
        },
    }
    r = client.post("/api/movimiento", json=payload, headers={"X-Admin-Key": "s3cret"})
    assert r.status_code == 422

    data = client.get("/api/all").json()
    assert data["movimientos"] == []
    assert data["historial"] == []


@pytest.mark.parametrize("campo,valor", [
    ("tipo", "regalo"),
    ("monto_usd", -100),
    ("monto_usd", 0),
    ("fecha", "10/01/2026"),
    ("persona", ""),
])
def test_movimiento_rechaza_datos_invalidos(client, campo, valor):
    payload = {
        "fecha": "2026-01-10T00:00", "persona": "Patico", "tipo": "aporte",
        "monto_usd": 100, "precio_cuota_dia": 1.0, "cuotas": 100, "monto_cop": 330000, "trm_dia": 3300,
    }
    payload[campo] = valor

    r = client.post("/api/movimiento", json=payload, headers={"X-Admin-Key": "s3cret"})
    assert r.status_code == 422
    assert client.get("/api/all").json()["movimientos"] == []


def test_error_de_validacion_devuelve_texto_legible(client):
    # El front muestra `detail` tal cual en el formulario: con la lista de errores de pydantic
    # el admin veía "[object Object]" y no qué campo estaba mal.
    payload = {
        "fecha": "2026-01-10T00:00", "persona": "Patico", "tipo": "aporte",
        "monto_usd": -5, "precio_cuota_dia": 1.0, "cuotas": 100, "monto_cop": 330000, "trm_dia": 3300,
    }
    r = client.post("/api/movimiento", json=payload, headers={"X-Admin-Key": "s3cret"})
    assert r.status_code == 422
    detail = r.json()["detail"]
    assert isinstance(detail, str)
    assert "monto_usd" in detail


def test_participante_rechaza_accion_desconocida(client):
    r = client.post("/api/participante", headers={"X-Admin-Key": "s3cret"},
                    json={"fecha": "2026-01-10T00:00", "nombre": "Patico", "accion": "borrar"})
    assert r.status_code == 422


def test_retiro_total_deja_el_fondo_en_cero(client):
    client.post("/api/movimiento", headers={"X-Admin-Key": "s3cret"}, json={
        "fecha": "2026-01-09T00:00", "persona": "Patico", "tipo": "aporte",
        "monto_usd": 100, "precio_cuota_dia": 1.0, "cuotas": 100, "monto_cop": 330000, "trm_dia": 3300,
    })

    payload = {
        "fecha": "2026-01-10T00:00", "persona": "Patico", "tipo": "retiro",
        "monto_usd": 100, "precio_cuota_dia": 1.0, "cuotas": -100, "monto_cop": 330000, "trm_dia": 3300,
        "fondo": {
            "fecha": "2026-01-10T00:00", "valor_total_usd": 0, "precio_cuota_usd": 1.0,
            "cuotas_en_circulacion": 0, "trm": 3300,
        },
    }
    r = client.post("/api/movimiento", json=payload, headers={"X-Admin-Key": "s3cret"})
    assert r.status_code == 201


def test_import_normaliza_tipo_y_rechaza_valores_desconocidos(client):
    ok = _xlsx_bytes({"movimientos": [
        ["fecha", "persona", "tipo", "monto", "precio_cuota_dia", "cuotas", "monto_cop", "trm_dia"],
        ["2026-01-10", "Patico", " Aporte ", 100, 1.0, 100, 330000, 3300],
    ]})
    r = client.post("/api/import", headers={"X-Admin-Key": "s3cret"},
                    files={"file": ("data.xlsx", ok, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200
    assert client.get("/api/all").json()["movimientos"][0]["tipo"] == "aporte"

    malo = _xlsx_bytes({"movimientos": [
        ["fecha", "persona", "tipo", "monto", "precio_cuota_dia", "cuotas", "monto_cop", "trm_dia"],
        ["2026-01-10", "Patico", "regalo", 100, 1.0, 100, 330000, 3300],
    ]})
    r = client.post("/api/import", headers={"X-Admin-Key": "s3cret"},
                    files={"file": ("data.xlsx", malo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 400
    assert "fila 2" in r.json()["detail"]

    # El import inválido no debe haber borrado lo que ya estaba.
    assert len(client.get("/api/all").json()["movimientos"]) == 1


def test_retiro_mayor_al_saldo_se_rechaza(client):
    aporte = {
        "fecha": "2026-01-10T00:00", "persona": "Patico", "tipo": "aporte",
        "monto_usd": 100, "precio_cuota_dia": 1.0, "cuotas": 100, "monto_cop": 330000, "trm_dia": 3300,
    }
    client.post("/api/movimiento", json=aporte, headers={"X-Admin-Key": "s3cret"})

    # Un cero de más en el retiro: dejaría a Patico con -900 cuotas y sin forma de deshacerlo.
    retiro = {**aporte, "tipo": "retiro", "monto_usd": 1000, "cuotas": -1000}
    r = client.post("/api/movimiento", json=retiro, headers={"X-Admin-Key": "s3cret"})
    assert r.status_code == 400

    assert len(client.get("/api/all").json()["movimientos"]) == 1


def test_retiro_total_pasa(client):
    aporte = {
        "fecha": "2026-01-10T00:00", "persona": "Patico", "tipo": "aporte",
        "monto_usd": 100, "precio_cuota_dia": 1.0, "cuotas": 100, "monto_cop": 330000, "trm_dia": 3300,
    }
    client.post("/api/movimiento", json=aporte, headers={"X-Admin-Key": "s3cret"})

    retiro = {**aporte, "tipo": "retiro", "cuotas": -100}
    r = client.post("/api/movimiento", json=retiro, headers={"X-Admin-Key": "s3cret"})
    assert r.status_code == 201


def test_auth_bloquea_despues_de_muchos_intentos(client):
    for _ in range(10):
        assert client.post("/api/auth/verify", headers={"X-Admin-Key": "no"}).status_code == 401

    r = client.post("/api/auth/verify", headers={"X-Admin-Key": "no"})
    assert r.status_code == 429
    # Bloqueo por origen, no por clave: la correcta tampoco pasa hasta que expire la ventana.
    assert client.post("/api/auth/verify", headers={"X-Admin-Key": "s3cret"}).status_code == 429


def test_export_reimportado_deja_la_db_igual(client):
    # El export es el único backup que sale de la máquina: tiene que poder volver a entrar
    # tal cual, incluida la validación de columnas requeridas del import.
    client.post("/api/movimiento", headers={"X-Admin-Key": "s3cret"}, json={
        "fecha": "2026-01-10T00:00", "persona": "Patico", "tipo": "aporte",
        "monto_usd": 100, "precio_cuota_dia": 1.0, "cuotas": 100, "monto_cop": 330000, "trm_dia": 3300,
        "fondo": {
            "fecha": "2026-01-10T00:00", "valor_total_usd": 100, "precio_cuota_usd": 1.0,
            "cuotas_en_circulacion": 100, "trm": 3300,
        },
    })
    client.post("/api/participante", headers={"X-Admin-Key": "s3cret"},
                json={"fecha": "2026-01-10T00:00", "nombre": "Patico", "accion": "agregar"})

    antes = client.get("/api/all").json()
    export = client.get("/api/export").content

    r = client.post("/api/import", headers={"X-Admin-Key": "s3cret"},
                    files={"file": ("fondi-export.xlsx", export, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200
    assert client.get("/api/all").json() == antes


def test_import_rechaza_archivo_gigante(client):
    grande = b"x" * (6 * 1024 * 1024)
    r = client.post("/api/import", headers={"X-Admin-Key": "s3cret"},
                    files={"file": ("data.xlsx", grande, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 413


def test_import_rechaza_hoja_sin_columnas_requeridas(client):
    # Header equivocado: sin este chequeo cada fila entraba con monto/cuotas en 0 y el
    # import destructivo dejaba la DB con movimientos que no significan nada.
    content = _xlsx_bytes({"movimientos": [
        ["fecha", "persona", "tipo", "valor", "cuotitas"],
        ["2026-01-10", "Patico", "aporte", 100, 100],
    ]})
    r = client.post("/api/import", headers={"X-Admin-Key": "s3cret"},
                    files={"file": ("data.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 400
    assert "monto" in r.json()["detail"]


def test_get_all_ordena_por_fecha(client):
    # El import respeta el orden de filas del archivo; /api/all tiene que devolverlas
    # ordenadas igual, porque el front resuelve "última acción gana" con ese orden.
    content = _xlsx_bytes({"participantes_config": [
        ["fecha", "nombre", "accion"],
        ["2026-03-01", "Patico", "quitar"],
        ["2026-01-01", "Patico", "agregar"],
        ["2026-06-01", "Patico", "agregar"],
    ]})
    client.post("/api/import", headers={"X-Admin-Key": "s3cret"},
                files={"file": ("data.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})

    fechas = [p["fecha"] for p in client.get("/api/all").json()["participantes_config"]]
    assert fechas == sorted(fechas)


def test_manda_security_headers(client):
    r = client.get("/api/health")
    assert "default-src 'self'" in r.headers["content-security-policy"]
    assert r.headers["x-content-type-options"] == "nosniff"


@pytest.mark.parametrize("literal", ["Infinity", "-Infinity", "NaN"])
def test_movimiento_rechaza_montos_no_finitos(client, literal):
    # json.loads acepta estos literales, `inf` pasa cualquier gt/ge y `nan` no tiene constraint:
    # se guardaban con 201 y después /api/all no podía serializarse nunca más (500 permanente).
    body = (
        '{"fecha":"2026-01-10T00:00","persona":"Patico","tipo":"aporte","monto_usd":%s,'
        '"precio_cuota_dia":1.0,"cuotas":100,"monto_cop":330000,"trm_dia":3300}' % literal
    )
    r = client.post("/api/movimiento", content=body,
                    headers={"X-Admin-Key": "s3cret", "Content-Type": "application/json"})
    assert r.status_code == 422
    assert client.get("/api/all").status_code == 200


def test_movimiento_rechaza_cuotas_no_finitas(client):
    body = (
        '{"fecha":"2026-01-10T00:00","persona":"Patico","tipo":"aporte","monto_usd":100,'
        '"precio_cuota_dia":1.0,"cuotas":NaN,"monto_cop":330000,"trm_dia":3300}'
    )
    r = client.post("/api/movimiento", content=body,
                    headers={"X-Admin-Key": "s3cret", "Content-Type": "application/json"})
    assert r.status_code == 422


def test_fondo_rechaza_valores_no_finitos(client):
    body = (
        '{"fecha":"2026-01-10T00:00","valor_total_usd":Infinity,"precio_cuota_usd":1.0,'
        '"cuotas_en_circulacion":100,"trm":3300}'
    )
    r = client.post("/api/fondo", content=body,
                    headers={"X-Admin-Key": "s3cret", "Content-Type": "application/json"})
    assert r.status_code == 422


def test_import_rechaza_celdas_no_finitas(client):
    # float("inf") funciona: una celda con ese texto entraba a la DB por la vía del import.
    content = _xlsx_bytes({"movimientos": [
        ["fecha", "persona", "tipo", "monto", "cuotas"],
        ["2026-01-10", "Patico", "aporte", "inf", 100],
    ]})
    r = client.post("/api/import", headers={"X-Admin-Key": "s3cret"},
                    files={"file": ("data.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 400
    assert "finito" in r.json()["detail"]
    assert client.get("/api/all").json()["movimientos"] == []


def test_rate_limit_usa_forwarded_for_solo_con_trust_proxy(tmp_path, monkeypatch):
    # Detrás de un proxy todos comparten IP y un intento fallido bloqueaba a todos. Con
    # TRUST_PROXY el contador es por cliente real; sin él el header se ignora, porque si no
    # cualquiera se saltaría el límite mandando un X-Forwarded-For distinto en cada intento.
    monkeypatch.setenv("DB_PATH", str(tmp_path / "test.db"))
    monkeypatch.setenv("ADMIN_PASSWORD", "s3cret")
    monkeypatch.setenv("TRUST_PROXY", "1")

    from app import db as db_module
    from app import main as main_module

    importlib.reload(db_module)
    importlib.reload(main_module)

    with TestClient(main_module.app) as c:
        for _ in range(10):
            c.post("/api/auth/verify", headers={"X-Admin-Key": "no", "X-Forwarded-For": "10.0.0.1"})

        bloqueado = c.post("/api/auth/verify", headers={"X-Admin-Key": "no", "X-Forwarded-For": "10.0.0.1"})
        assert bloqueado.status_code == 429

        otro = c.post("/api/auth/verify", headers={"X-Admin-Key": "s3cret", "X-Forwarded-For": "10.0.0.2"})
        assert otro.status_code == 200


def test_rate_limit_ignora_forwarded_for_sin_trust_proxy(client):
    for i in range(10):
        client.post("/api/auth/verify", headers={"X-Admin-Key": "no", "X-Forwarded-For": f"10.0.0.{i}"})

    r = client.post("/api/auth/verify", headers={"X-Admin-Key": "no", "X-Forwarded-For": "10.0.0.99"})
    assert r.status_code == 429


def test_import_hace_backup_de_la_db(client, tmp_path):
    client.post("/api/participante", headers={"X-Admin-Key": "s3cret"},
                json={"fecha": "2026-01-01T00:00", "nombre": "Viejo", "accion": "agregar"})

    content = _xlsx_bytes({"participantes_config": [
        ["fecha", "nombre", "accion"],
        ["2026-01-10", "Nuevo", "agregar"],
    ]})
    r = client.post("/api/import", headers={"X-Admin-Key": "s3cret"},
                    files={"file": ("data.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200

    backups = list(tmp_path.glob("test.db.*.bak"))
    assert len(backups) == 1

    viejo = sqlite3.connect(backups[0])
    nombres = [row[0] for row in viejo.execute("SELECT nombre FROM participantes_config")]
    viejo.close()
    assert nombres == ["Viejo"]
